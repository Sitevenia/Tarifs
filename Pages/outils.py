# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import math
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import re

# --- FONCTIONS UTILITAIRES COMMUNES ---
def clean_col_names(df):
    cols = df.columns
    new_cols = []
    for col in cols:
        new_col = str(col).strip().lower(); new_col = re.sub(r"[ ']", '_', new_col); new_col = re.sub(r'[éèê]', 'e', new_col)
        new_col = new_col.replace('fournisseur_principal', 'fournisseur_principal')
        new_col = new_col.replace('nom_fournisseur', 'nom_fournisseur')
        new_cols.append(new_col)
    df.columns = new_cols
    return df

def rechercher_produits(df, references):
    df_recherche = df.copy()
    df_recherche['reference_article'] = df_recherche['reference_article'].astype(str)
    df_recherche['af_reffourniss'] = df_recherche['af_reffourniss'].astype(str)
    df_recherche['reference_mbt'] = df_recherche['reference_mbt'].astype(str)
    references_str = [str(r) for r in references]
    produits_selectionnes = df_recherche[df_recherche['reference_article'].isin(references_str) | df_recherche['af_reffourniss'].isin(references_str) | df_recherche['reference_mbt'].isin(references_str)].copy()
    return produits_selectionnes

# --- FONCTIONS POUR L'ONGLET 1: CALCULATEUR STANDARD ---
def calculer_prix_de_vente_std(produits_df_modifies, marge):
    df_final = produits_df_modifies.copy()
    df_final['prix_d_achat'] = pd.to_numeric(df_final['prix_d_achat'], errors='coerce')
    df_final['dernier_prix_d_achat'] = pd.to_numeric(df_final['dernier_prix_d_achat'], errors='coerce')
    diviseur = 1 - (marge / 100)
    if diviseur <= 0: diviseur = 0.001
    df_final['PV HT (base P.A.)'] = df_final['prix_d_achat'] / diviseur
    df_final['PV HT (base D.P.A.)'] = df_final['dernier_prix_d_achat'] / diviseur
    return df_final

def calculer_totaux_et_comparaison_std(df_final, arrondi, taux_tva=20.0):
    df_final['quantite'] = pd.to_numeric(df_final['quantite'], errors='coerce').fillna(1)
    total_ht_pa = (df_final['PV HT (base P.A.)'] * df_final['quantite']).sum()
    total_ht_dpa = (df_final['PV HT (base D.P.A.)'] * df_final['quantite']).sum()
    if arrondi:
        total_ht_pa = math.ceil(total_ht_pa) - 0.10
        total_ht_dpa = math.ceil(total_ht_dpa) - 0.10
    total_actuel_ttl = 0
    if 'prix_de_vente' in df_final.columns:
        df_final['prix_de_vente'] = pd.to_numeric(df_final['prix_de_vente'], errors='coerce')
        total_actuel_ttl = (df_final['prix_de_vente'] * df_final['quantite']).sum()
    total_actuel_mbt = 0
    if 'prix_de_vente_mbt' in df_final.columns:
        df_final['prix_de_vente_mbt'] = pd.to_numeric(df_final['prix_de_vente_mbt'], errors='coerce')
        total_actuel_mbt = (df_final['prix_de_vente_mbt'] * df_final['quantite']).sum()
    tva_multiplier = 1 + (taux_tva / 100)
    resume = {
        "total_ht_pa": total_ht_pa, "total_ttc_pa": total_ht_pa * tva_multiplier,
        "total_ht_dpa": total_ht_dpa, "total_ttc_dpa": total_ht_dpa * tva_multiplier,
        "total_actuel_ttl": total_actuel_ttl, "total_actuel_mbt": total_actuel_mbt,
        "total_actuel_ttl_ttc": total_actuel_ttl * tva_multiplier,
        "total_actuel_mbt_ttc": total_actuel_mbt * tva_multiplier
    }
    return resume

def exporter_vers_excel_buffer_std(produits_df, resume, entreprise, marge, arrondi):
    wb = Workbook(); ws = wb.active; ws.title = "Calcul Prix de Vente"
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid"); summary_font = Font(bold=True)
    ws.merge_cells('A1:B1'); ws['A1'] = "Parametres du Calcul"; ws['A1'].font = summary_font
    ws['A2'] = "Entreprise:"; ws['B2'] = entreprise; ws['A3'] = "Marge appliquee:"; ws['B3'] = f"{marge}%"; ws['A4'] = "Tarif arrondi:"; ws['B4'] = "Oui" if arrondi else "Non"
    start_row = 6
    ws.cell(row=start_row -1, column=1, value="Détail de la Proposition Tarifaire").font = summary_font
    headers = ['Reference article', 'Designation article', 'Quantite', 'Prix d\'achat', 'PV HT unitaire (base P.A.)', 'Dernier Prix d\'achat', 'PV HT unitaire (base D.P.A.)']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    cols_a_afficher = ['reference_article', 'designation_article', 'quantite', 'prix_d_achat', 'PV HT (base P.A.)', 'dernier_prix_d_achat', 'PV HT (base D.P.A.)']
    for row_index, row_data in enumerate(produits_df[cols_a_afficher].itertuples(index=False), start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_num, value=cell_value)
            if col_num >= 4: cell.number_format = '#,##0.00'
    summary_row_start = start_row + len(produits_df) + 2
    ws.cell(row=summary_row_start, column=4, value="Total HT:").alignment = Alignment(horizontal='right'); ws.cell(row=summary_row_start + 1, column=4, value="Total TTC:").alignment = Alignment(horizontal='right')
    ws.cell(row=summary_row_start, column=6, value="Total HT:").alignment = Alignment(horizontal='right'); ws.cell(row=summary_row_start + 1, column=6, value="Total TTC:").alignment = Alignment(horizontal='right')
    ws.cell(row=summary_row_start, column=5, value=resume['total_ht_pa']).number_format = '#,##0.00'; ws.cell(row=summary_row_start, column=5).font = summary_font
    ws.cell(row=summary_row_start + 1, column=5, value=resume['total_ttc_pa']).number_format = '#,##0.00'; ws.cell(row=summary_row_start + 1, column=5).font = summary_font
    ws.cell(row=summary_row_start, column=7, value=resume['total_ht_dpa']).number_format = '#,##0.00'; ws.cell(row=summary_row_start, column=7).font = summary_font
    ws.cell(row=summary_row_start + 1, column=7, value=resume['total_ttc_dpa']).number_format = '#,##0.00'; ws.cell(row=summary_row_start + 1, column=7).font = summary_font
    comp_summary_row = summary_row_start + 4 + len(produits_df) + 1
    ws.cell(row=comp_summary_row, column=4, value="Total Actuel Tetenal:").alignment = Alignment(horizontal='right'); ws.cell(row=comp_summary_row, column=6, value="Total Actuel MB TECH:").alignment = Alignment(horizontal='right')
    ws.cell(row=comp_summary_row, column=5, value=resume['total_actuel_ttl']).number_format = '#,##0.00'; ws.cell(row=comp_summary_row, column=5).font = summary_font
    ws.cell(row=comp_summary_row, column=7, value=resume['total_actuel_mbt']).number_format = '#,##0.00'; ws.cell(row=comp_summary_row, column=7).font = summary_font
    analysis_start_row = comp_summary_row + 3
    headers_analysis = ['Cible de la Proposition', 'Total Proposé (€)', 'Total Actuel (€)', 'Remise (€)', 'Remise (%)']
    ws.cell(row=analysis_start_row - 1, column=1, value="Analyse des Remises vs. Tarif Actuel TETENAL").font = summary_font
    for col_num, header_title in enumerate(headers_analysis, 1):
        cell = ws.cell(row=analysis_start_row, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    ttl_data = [("HT (base P.A.)", resume['total_ht_pa'], resume['total_actuel_ttl']), ("TTC (base P.A.)", resume['total_ttc_pa'], resume['total_actuel_ttl_ttc']), ("HT (base D.P.A.)", resume['total_ht_dpa'], resume['total_actuel_ttl']), ("TTC (base D.P.A.)", resume['total_ttc_dpa'], resume['total_actuel_ttl_ttc'])]
    current_row = analysis_start_row + 1
    for label, prop_total, base_total in ttl_data:
        remise_val = base_total - prop_total; remise_pct = (remise_val / base_total) if base_total > 0 else 0
        ws.cell(row=current_row, column=1, value=label); ws.cell(row=current_row, column=2, value=prop_total).number_format = '#,##0.00€'; ws.cell(row=current_row, column=3, value=base_total).number_format = '#,##0.00€'; ws.cell(row=current_row, column=4, value=remise_val).number_format = '#,##0.00€'; ws.cell(row=current_row, column=5, value=remise_pct).number_format = '0.00%'
        current_row += 1
    mbt_start_row = current_row + 2
    ws.cell(row=mbt_start_row - 1, column=1, value="Analyse des Remises vs. Tarif Actuel MB TECH").font = summary_font
    for col_num, header_title in enumerate(headers_analysis, 1):
        cell = ws.cell(row=mbt_start_row, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    mbt_data = [("HT (base P.A.)", resume['total_ht_pa'], resume['total_actuel_mbt']), ("TTC (base P.A.)", resume['total_ttc_pa'], resume['total_actuel_mbt_ttc']), ("HT (base D.P.A.)", resume['total_ht_dpa'], resume['total_actuel_mbt']), ("TTC (base D.P.A.)", resume['total_ttc_dpa'], resume['total_actuel_mbt_ttc'])]
    current_row = mbt_start_row + 1
    for label, prop_total, base_total in mbt_data:
        remise_val = base_total - prop_total; remise_pct = (remise_val / base_total) if base_total > 0 else 0
        ws.cell(row=current_row, column=1, value=label); ws.cell(row=current_row, column=2, value=prop_total).number_format = '#,##0.00€'; ws.cell(row=current_row, column=3, value=base_total).number_format = '#,##0.00€'; ws.cell(row=current_row, column=4, value=remise_val).number_format = '#,##0.00€'; ws.cell(row=current_row, column=5, value=remise_pct).number_format = '0.00%'
        current_row += 1
    for col in ws.columns:
        max_length = 0; column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value: max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer

# --- FONCTIONS POUR L'ONGLET 2: CALCULATEUR DE PROMOTIONS ---
def calculer_et_analyser_promo(df_promo, taux_tva=20.0):
    methode = st.session_state.promo_methode
    tva_multiplier = 1 + (taux_tva / 100)
    df_promo['prix_achat_promo'] = pd.to_numeric(df_promo['prix_achat_promo'], errors='coerce').fillna(0)
    df_promo['prix_de_vente'] = pd.to_numeric(df_promo.get('prix_de_vente'), errors='coerce').fillna(0)
    if methode == "Appliquer une marge":
        marge = st.session_state.promo_marge
        arrondi = st.session_state.promo_arrondi
        diviseur = 1 - (marge / 100)
        if diviseur <= 0: diviseur = 0.001
        df_promo['pv_promo_ht'] = df_promo['prix_achat_promo'] / diviseur
        if arrondi:
            df_promo['pv_promo_ht'] = df_promo['pv_promo_ht'].apply(lambda x: math.ceil(x) - 0.10 if pd.notna(x) else x)
        df_promo['marge_degagee_%'] = marge / 100
    elif methode == "Définir un prix de vente cible":
        df_promo['pv_promo_ht'] = pd.to_numeric(df_promo['pv_promo_ht'], errors='coerce').fillna(0)
        df_promo['marge_degagee_%'] = (1 - (df_promo['prix_achat_promo'] / df_promo['pv_promo_ht']))
        df_promo['marge_degagee_%'] = df_promo['marge_degagee_%'].fillna(0).replace([float('inf'), -float('inf')], 0)
    df_promo['pv_promo_ttc'] = df_promo['pv_promo_ht'] * tva_multiplier
    df_promo['pv_actuel_ttc'] = df_promo['prix_de_vente'] * tva_multiplier
    df_promo['remise_valeur_ttc_€'] = df_promo['pv_actuel_ttc'] - df_promo['pv_promo_ttc']
    df_promo['remise_valeur_ht_€'] = df_promo['prix_de_vente'] - df_promo['pv_promo_ht']
    df_promo['remise_%'] = (df_promo['remise_valeur_ht_€'] / df_promo['prix_de_vente'])
    df_promo['remise_%'] = df_promo['remise_%'].fillna(0).replace([float('inf'), -float('inf')], 0)
    return df_promo

def exporter_promo_vers_excel(df_resultats, nom_promo, methode):
    wb = Workbook(); ws = wb.active; ws.title = "Rapport de Promotion"
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid"); summary_font = Font(bold=True)
    ws.merge_cells('A1:B1'); ws['A1'] = "Parametres de la Promotion"; ws['A1'].font = summary_font
    ws['A2'] = "Nom de la promotion:"; ws['B2'] = nom_promo
    ws['A3'] = "Methode de calcul:"; ws['B3'] = methode
    start_row = 5
    ws.cell(row=start_row -1, column=1, value="Détail de la Promotion").font = summary_font
    headers = ['Reference article', 'Designation article', 'Quantite', 'Prix Achat Actuel', 'Prix Achat Promo', 'Prix Vente Actuel HT', 'Prix Vente Promo HT', 'Prix Vente Promo TTC', 'Marge Dégagée (%)', 'Remise (€) HT', 'Remise (€) TTC', 'Remise (%)']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    cols_a_afficher = ['reference_article', 'designation_article', 'quantite', 'prix_d_achat', 'prix_achat_promo', 'prix_de_vente', 'pv_promo_ht', 'pv_promo_ttc', 'marge_degagee_%', 'remise_valeur_ht_€', 'remise_valeur_ttc_€', 'remise_%']
    for row_index, row_data in enumerate(df_resultats[cols_a_afficher].itertuples(index=False), start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_num, value=cell_value)
            if col_num in [4, 5, 6, 7, 8, 10, 11]: cell.number_format = '#,##0.00€'
            if col_num in [9, 12]: cell.number_format = '0.00%'
    for col in ws.columns:
        max_length = 0; column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value: max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer

# --- FONCTIONS POUR L'ONGLET 3: COMPARATEUR TTL-MBT ---
def run_comparison(df, filters, analyze_all):
    df_filtered = df.copy()
    if not analyze_all:
        if filters['references']:
            df_filtered = df_filtered[df_filtered['reference_article'].isin(filters['references'])]
        if 'nom_fournisseur' in df_filtered.columns and filters['nom_fournisseur'] != "Tous":
            df_filtered = df_filtered[df_filtered['nom_fournisseur'] == filters['nom_fournisseur']]
        if 'code_famille' in df_filtered.columns and filters['code_famille'] != "Tous":
            df_filtered = df_filtered[df_filtered['code_famille'] == filters['code_famille']]
    for col in ['prix_de_vente', 'prix_de_vente_mbt', 'dernier_prix_d_achat']:
        if col in df_filtered.columns:
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce')
        else:
            df_filtered[col] = np.nan
    df_comp = df_filtered.dropna(subset=['prix_de_vente', 'prix_de_vente_mbt'])
    df_comp = df_comp[(df_comp['prix_de_vente'] > 0) & (df_comp['prix_de_vente_mbt'] > 0)]
    if df_comp.empty:
        return pd.DataFrame(), pd.DataFrame()
    df_comp['ecart_€'] = df_comp['prix_de_vente_mbt'] - df_comp['prix_de_vente']
    df_comp = df_comp[df_comp['ecart_€'] != 0].copy()
    df_comp['ecart_%'] = df_comp['ecart_€'] / df_comp['prix_de_vente']
    df_comp['marge_ttl_%'] = (1 - (df_comp['dernier_prix_d_achat'] / df_comp['prix_de_vente']))
    df_comp['marge_mbt_%'] = (1 - (df_comp['dernier_prix_d_achat'] / df_comp['prix_de_vente_mbt']))
    alertes = []
    for _, row in df_comp.iterrows():
        alerte = []
        if 'tarif_vente_-_valeur_remise' in row and pd.notna(row['tarif_vente_-_valeur_remise']) and row['tarif_vente_-_valeur_remise'] != 0:
            alerte.append("Remise TTL")
        if 'remise_mbt' in row and pd.notna(row['remise_mbt']) and "Gamme" in str(row['remise_mbt']):
            alerte.append("Remise MBT")
        alertes.append(", ".join(alerte))
    df_comp['alertes'] = alertes
    for col in ['ecart_%', 'marge_ttl_%', 'marge_mbt_%']:
        df_comp[col] = df_comp[col].replace([np.inf, -np.inf], np.nan)
    df_moins_cher = df_comp[df_comp['ecart_€'] < 0].copy()
    df_plus_cher = df_comp[df_comp['ecart_€'] > 0].copy()
    if analyze_all and 'nom_fournisseur' in df_comp.columns:
        df_moins_cher = df_moins_cher.sort_values(by=['nom_fournisseur', 'reference_article'])
        df_plus_cher = df_plus_cher.sort_values(by=['nom_fournisseur', 'reference_article'])
    return df_moins_cher, df_plus_cher

def run_full_list(df, filters, analyze_all):
    """Génère la liste complète des produits sélectionnés avec les colonnes demandées."""
    df_filtered = df.copy()
    if not analyze_all:
        if filters['references']:
            df_filtered = df_filtered[df_filtered['reference_article'].isin(filters['references'])]
        if 'nom_fournisseur' in df_filtered.columns and filters['nom_fournisseur'] != "Tous":
            df_filtered = df_filtered[df_filtered['nom_fournisseur'] == filters['nom_fournisseur']]
        if 'code_famille' in df_filtered.columns and filters['code_famille'] != "Tous":
            df_filtered = df_filtered[df_filtered['code_famille'] == filters['code_famille']]
    
    # Création du DataFrame de sortie avec renommage
    res = df_filtered[['reference_article', 'reference_mbt', 'designation_article', 'prix_de_vente', 'prix_de_vente_mbt']].copy()
    
    # Gestion des produits non référencés chez MBT
    res['reference_mbt'] = res['reference_mbt'].astype(str).replace(['nan', 'None', ''], np.nan)
    res['reference_mbt'] = res['reference_mbt'].fillna("Produit non référencé")
    
    # Conversion prix en numérique
    for col in ['prix_de_vente', 'prix_de_vente_mbt']:
        res[col] = pd.to_numeric(res[col], errors='coerce')
        
    return res

def exporter_comparaison_vers_excel(df_moins_cher, df_plus_cher, df_full, filters, analyze_all):
    wb = Workbook(); wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid"); summary_font = Font(bold=True)
    
    def setup_sheet(sheet, title, df, is_full_list=False):
        sheet.merge_cells('A1:B1'); sheet['A1'] = "Parametres de la Comparaison"; sheet['A1'].font = summary_font
        if analyze_all:
            sheet['A2'] = "Mode d'analyse:"; sheet['B2'] = "Analyse Globale de tous les écarts"
        else:
            sheet['A2'] = "Filtres appliqués:"; sheet['B2'] = f"Fournisseur: {filters['nom_fournisseur']}, Famille: {filters['code_famille']}"
        
        # Définition des en-têtes selon le type d'onglet
        if is_full_list:
            headers = ['Référence TTL', 'Référence MBT', 'Désignation', 'Prix de vente TTL', 'Prix de vente MBT']
            cols = ['reference_article', 'reference_mbt', 'designation_article', 'prix_de_vente', 'prix_de_vente_mbt']
        else:
            base_headers = ['Reference article', 'Designation article', 'PV TTL', 'PV MBT', 'Ecart (€)', 'Ecart (%)', 'Dernier PA', 'Marge TTL (%)', 'Marge MBT (%)', 'Alertes']
            base_cols = ['reference_article', 'designation_article', 'prix_de_vente', 'prix_de_vente_mbt', 'ecart_€', 'ecart_%', 'dernier_prix_d_achat', 'marge_ttl_%', 'marge_mbt_%', 'alertes']
            sample_df = df_moins_cher if not df_moins_cher.empty else df_plus_cher
            headers, cols = base_headers[:], base_cols[:]
            if not sample_df.empty:
                if 'nom_fournisseur' in sample_df.columns:
                    headers.insert(0, 'Nom Fournisseur'); cols.insert(0, 'nom_fournisseur')
                if 'code_famille' in sample_df.columns:
                    headers.insert(1, 'Code Famille'); cols.insert(1, 'code_famille')
        
        sheet.cell(row=4, column=1, value=title).font = summary_font
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        
        if df.empty:
            sheet.cell(row=6, column=1, value="Aucun produit ne correspond à cette catégorie.")
        else:
            for row_index, row_data in enumerate(df[cols].itertuples(index=False), 6):
                for col_num, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_index, column=col_num, value=cell_value)
        
        # Formatage des cellules
        for row in sheet.iter_rows(min_row=6, max_row=5 + len(df), min_col=1, max_col=len(headers)):
            for cell in row:
                header = headers[cell.column - 1]
                if any(x in header for x in ['PV', 'Prix de vente', 'Ecart (€)', 'PA']):
                    cell.number_format = '#,##0.00€'
                if '%' in header:
                    cell.number_format = '0.00%'
                    
        for col in sheet.columns:
            max_length = 0; column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    # Création des 3 onglets
    setup_sheet(wb.create_sheet(title="Liste Complète TTL-MBT"), "📋 Récapitulatif Complet des Tarifs", df_full, is_full_list=True)
    setup_sheet(wb.create_sheet(title="MBT Moins Cher"), "✅ MBT est Moins Cher que TTL", df_moins_cher)
    setup_sheet(wb.create_sheet(title="MBT Plus Cher"), "❌ MBT est Plus Cher que TTL", df_plus_cher)
    
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer

# --- FONCTIONS POUR L'ONGLET 4: VÉRIFICATEUR DE MARGES ---
def run_margin_check(df, threshold, filters, analyze_all):
    df_filtered = df.copy()
    if not analyze_all:
        if filters['references']:
            df_filtered = df_filtered[df_filtered['reference_article'].isin(filters['references'])]
        if 'nom_fournisseur' in df_filtered.columns and filters['nom_fournisseur'] != "Tous":
            df_filtered = df_filtered[df_filtered['nom_fournisseur'] == filters['nom_fournisseur']]
        if 'code_famille' in df_filtered.columns and filters['code_famille'] != "Tous":
            df_filtered = df_filtered[df_filtered['code_famille'] == filters['code_famille']]
    for col in ['prix_de_vente', 'prix_de_vente_mbt', 'dernier_prix_d_achat']:
        if col in df_filtered.columns:
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce')
        else:
            df_filtered[col] = np.nan
    df_check = df_filtered.dropna(subset=['dernier_prix_d_achat'])
    df_check = df_check[df_check['dernier_prix_d_achat'] > 0]
    if df_check.empty:
        return pd.DataFrame(), pd.DataFrame()
    df_check['marge_ttl_%'] = (1 - (df_check['dernier_prix_d_achat'] / df_check['prix_de_vente'])).replace([np.inf, -np.inf], np.nan)
    df_check['marge_mbt_%'] = (1 - (df_check['dernier_prix_d_achat'] / df_check['prix_de_vente_mbt'])).replace([np.inf, -np.inf], np.nan)
    threshold_decimal = threshold / 100
    alerts_ttl = df_check[df_check['marge_ttl_%'] < threshold_decimal].copy()
    alerts_mbt = df_check[df_check['marge_mbt_%'] < threshold_decimal].copy()
    if not alerts_ttl.empty:
        alerts_ttl['alerte_remise'] = alerts_ttl.apply(
            lambda row: "Attention remise" if 'tarif_vente_-_valeur_remise' in row and pd.notna(row['tarif_vente_-_valeur_remise']) and row['tarif_vente_-_valeur_remise'] != 0 else "",
            axis=1
        )
    if not alerts_mbt.empty:
        alerts_mbt['alerte_remise'] = alerts_mbt.apply(
            lambda row: "Attention remise" if 'remise_mbt' in row and pd.notna(row['remise_mbt']) and "Gamme" in str(row['remise_mbt']) else "",
            axis=1
        )
    return alerts_ttl, alerts_mbt

def exporter_marges_vers_excel(df_ttl, df_mbt, threshold):
    wb = Workbook(); wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid"); summary_font = Font(bold=True)
    def setup_sheet(sheet, title, df, margin_col_name, pv_col_name):
        sheet['A1'] = f"Analyse des Marges - {title}"; sheet['A1'].font = summary_font
        sheet['A2'] = f"Seuil d'alerte: Marge < {threshold}%"
        headers = ['Reference article', 'Designation article', 'Dernier Prix d\'Achat', 'Prix de Vente', 'Marge Actuelle (%)', 'Alerte Remise']
        cols_to_display = ['reference_article', 'designation_article', 'dernier_prix_d_achat', pv_col_name, margin_col_name, 'alerte_remise']
        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        if df.empty:
            sheet.cell(row=5, column=1, value="Aucun produit en alerte de marge.")
        else:
            for row_index, row_data in enumerate(df[cols_to_display].itertuples(index=False), 5):
                for col_num, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_index, column=col_num, value=cell_value)
        for row in sheet.iter_rows(min_row=5, max_row=4 + len(df), min_col=1, max_col=len(headers)):
            for cell in row:
                if headers[cell.column - 1] in ['Dernier Prix d\'Achat', 'Prix de Vente']:
                    cell.number_format = '#,##0.00€'
                if headers[cell.column - 1] == 'Marge Actuelle (%)':
                    cell.number_format = '0.00%'
        for col in sheet.columns:
            max_length = 0; column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2
    setup_sheet(wb.create_sheet("Alertes Marge TTL"), "Tetenal", df_ttl, 'marge_ttl_%', 'prix_de_vente')
    setup_sheet(wb.create_sheet("Alertes Marge MBT"), "MB TECH", df_mbt, 'marge_mbt_%', 'prix_de_vente_mbt')
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer

# --- FONCTIONS POUR L'ONGLET 5: NOMENCLATURES ---
def calculer_simulation_nomenclature(df_items, pv_cible, marge_cible, arrondi, taux_tva=20.0):
    df = df_items.copy()
    # Conversion PA et PV
    for col in ['prix_d_achat', 'prix_de_vente', 'prix_de_vente_mbt', 'quantite']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    df['total_pa_ligne'] = df['prix_d_achat'] * df['quantite']
    df['total_ttl_ligne'] = df['prix_de_vente'] * df['quantite']
    df['total_mbt_ligne'] = df['prix_de_vente_mbt'] * df['quantite']
    
    total_cout_pa_ht = df['total_pa_ligne'].sum()
    total_normal_ttl_ht = df['total_ttl_ligne'].sum()
    total_normal_mbt_ht = df['total_mbt_ligne'].sum()
    
    tva_multiplier = 1 + (taux_tva / 100)
    
    # Sim 1 : Prix Cible
    pv1_ht = pv_cible
    marge1 = (1 - (total_cout_pa_ht / pv1_ht)) if pv1_ht > 0 else 0
    pv1_ttc = pv1_ht * tva_multiplier
    
    # Sim 2 : Marge Cible
    diviseur = 1 - (marge_cible / 100)
    if diviseur <= 0: diviseur = 0.001
    pv2_ht = total_cout_pa_ht / diviseur
    if arrondi:
        pv2_ht = math.ceil(pv2_ht) - 0.10
    marge2 = marge_cible / 100
    pv2_ttc = pv2_ht * tva_multiplier
    
    return {
        'df': df,
        'cout_total_pa': total_cout_pa_ht,
        'total_normal_ttl': total_normal_ttl_ht,
        'total_normal_mbt': total_normal_mbt_ht,
        'sim_prix': {'pv_ht': pv1_ht, 'pv_ttc': pv1_ttc, 'marge': marge1},
        'sim_marge': {'pv_ht': pv2_ht, 'pv_ttc': pv2_ttc, 'marge': marge2}
    }

def exporter_nomenclature_comparative_excel(bundle_res, bundle_name):
    wb = Workbook(); ws = wb.active; ws.title = "Nomenclature Bundle"
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid"); summary_font = Font(bold=True)
    
    ws.merge_cells('A1:C1'); ws['A1'] = f"Nomenclature : {bundle_name}"; ws['A1'].font = Font(bold=True, size=14)
    
    start_row = 4
    # Headers incluant les prix individuels TTL et MBT
    headers = ['Référence article', 'Désignation article', 'Quantité', 'PA Unit.', 'PV TTL Unit.', 'PV MBT Unit.', 'Total PA Ligne']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    
    df_items = bundle_res['df']
    cols_to_map = ['reference_article', 'designation_article', 'quantite', 'prix_d_achat', 'prix_de_vente', 'prix_de_vente_mbt', 'total_pa_ligne']
    
    for row_index, row_data in enumerate(df_items[cols_to_map].itertuples(index=False), start_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_num, value=cell_value)
            if col_num >= 4: cell.number_format = '#,##0.00€'
            
    res_row = start_row + len(df_items) + 3
    ws.cell(row=res_row, column=1, value="RÉSULTATS DES SIMULATIONS ET VALEURS NORMALES").font = summary_font
    
    # En-têtes du tableau de simulation formatés
    headers_sim = ['', 'Simulation par Prix Cible', 'Simulation par Marge Cible']
    for col_idx, val in enumerate(headers_sim, 1):
        cell = ws.cell(row=res_row + 1, column=col_idx, value=val)
        if val != '':
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    labels = ["Coût Total Achat (HT)", "Prix de Vente Final (HT)", "Prix de Vente Final (TTC)", "Marge dégagée (%)"]
    s1, s2 = bundle_res['sim_prix'], bundle_res['sim_marge']
    
    data_rows = [
        (bundle_res['cout_total_pa'], bundle_res['cout_total_pa']),
        (s1['pv_ht'], s2['pv_ht']),
        (s1['pv_ttc'], s2['pv_ttc']),
        (s1['marge'], s2['marge'])
    ]
    
    for i, label in enumerate(labels):
        ws.cell(row=res_row + 2 + i, column=1, value=label).font = summary_font
        c_s1 = ws.cell(row=res_row + 2 + i, column=2, value=data_rows[i][0])
        c_s2 = ws.cell(row=res_row + 2 + i, column=3, value=data_rows[i][1])
        if "Marge" in label:
            c_s1.number_format = '0.00%'; c_s2.number_format = '0.00%'
        else:
            c_s1.number_format = '#,##0.00€'; c_s2.number_format = '#,##0.00€'

    # Ajout des totaux normaux Bundle
    norm_row = res_row + 7
    ws.cell(row=norm_row, column=1, value="Valeur Totale Bundle (Prix Normal)").font = summary_font
    ws.cell(row=norm_row + 1, column=1, value="Total Normal Tetenal (HT)").font = summary_font
    ws.cell(row=norm_row + 1, column=2, value=bundle_res['total_normal_ttl']).number_format = '#,##0.00€'
    ws.cell(row=norm_row + 2, column=1, value="Total Normal MB TECH (HT)").font = summary_font
    ws.cell(row=norm_row + 2, column=2, value=bundle_res['total_normal_mbt']).number_format = '#,##0.00€'

    for col in ws.columns:
        max_length = 0; column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value: max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
        
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer

# --- INTERFACE PRINCIPALE DE L'APPLICATION ---
st.set_page_config(page_title="Outil d'Analyse et de Calcul de Prix", layout="wide")
st.title("Outil d'Analyse et de Calcul de Prix")

if 'df_base_clean' not in st.session_state: st.session_state.df_base_clean = None

with st.sidebar:
    uploaded_file = st.file_uploader("1. Chargez votre fichier de base de données Excel", type=["xlsx", "xls"])
    if uploaded_file:
        df_base = pd.read_excel(uploaded_file, header=1)
        st.session_state.df_base_clean = clean_col_names(df_base)
        st.success("Fichier chargé !")
    st.markdown("---")
    if st.button("🔄 Réinitialiser l'application"):
        df_base = st.session_state.get('df_base_clean')
        st.session_state.clear()
        st.session_state.df_base_clean = df_base
        st.rerun()

tab1, tab2, tab3, tab4, tab5 = st.tabs(["Calculateur Standard", "Calculateur de Promotions", "Comparateur TTL-MBT", "Vérificateur de Marges", "Nomenclatures (Bundles)"])

with tab1:
    if st.session_state.df_base_clean is None: st.info("Veuillez charger un fichier Excel pour commencer.")
    else:
        df_base_clean = st.session_state.df_base_clean
        with st.expander("⚙️ Étape 1: Configuration et Sélection des Produits", expanded=True):
            st.text_input("Nom du calcul (pour le nom du fichier exporté)", key="std_nom_calcul", placeholder="Ex: Devis Client Dupont")
            col1, col2 = st.columns(2)
            with col1:
                entreprise = st.selectbox("2. Entreprise", ["Tetenal", "MB TECH"], key="std_entreprise"); marge = st.number_input("4. Marge (%)", min_value=0.0, max_value=99.9, value=30.0, step=0.5, format="%.2f", key="std_marge")
            with col2:
                all_references = sorted(pd.concat([df_base_clean['reference_article'].dropna().astype(str), df_base_clean['af_reffourniss'].dropna().astype(str), df_base_clean['reference_mbt'].dropna().astype(str)]).unique())
                references_selectionnees = st.multiselect("3. Produits", options=all_references, key="std_selection"); arrondi = st.checkbox("5. Appliquer l'arrondi final (total en ,90€)", key="std_arrondi")
            if st.button("Afficher les produits pour modification", key="std_btn_afficher"):
                if not references_selectionnees: st.warning("Veuillez selectionner au moins un produit.")
                else:
                    produits = rechercher_produits(df_base_clean, references_selectionnees).copy(); produits['quantite'] = 1
                    st.session_state.std_df_a_modifier = produits
                    if 'std_resume_final' in st.session_state: del st.session_state.std_resume_final
                    if 'std_produits_finaux' in st.session_state: del st.session_state.std_produits_finaux
        if 'std_df_a_modifier' in st.session_state:
            st.header("Étape 2: Modifiez les prix d'achat et les quantités")
            cols_a_editer = ['reference_article', 'designation_article', 'prix_d_achat', 'dernier_prix_d_achat', 'quantite']
            st.session_state.std_df_modifie = st.data_editor(st.session_state.std_df_a_modifier[cols_a_editer], key="std_editor_prix_achat", disabled=['reference_article', 'designation_article'], num_rows="dynamic")
            if st.button("✅ Calculer les Prix de Vente Finaux", key="std_btn_calculer"):
                df_complet_modifie = st.session_state.std_df_a_modifier.copy(); df_complet_modifie.update(st.session_state.std_df_modifie)
                df_avec_pv = calculer_prix_de_vente_std(df_complet_modifie, st.session_state.std_marge)
                resume = calculer_totaux_et_comparaison_std(df_avec_pv, st.session_state.std_arrondi)
                st.session_state.std_produits_finaux = df_avec_pv; st.session_state.std_resume_final = resume
        if 'std_produits_finaux' in st.session_state and 'std_resume_final' in st.session_state:
            st.header("Étape 3: Résultats Finaux")
            st.subheader("Prix de vente unitaires calculés")
            st.dataframe(st.session_state.std_produits_finaux[['reference_article', 'designation_article', 'quantite', 'PV HT (base P.A.)', 'PV HT (base D.P.A.)']])
            st.subheader("Totaux Finaux (tenant compte des quantités)")
            resume_final = st.session_state.std_resume_final; col1, col2 = st.columns(2)
            with col1:
                st.metric(label="Total HT Final (base P.A.)", value=f"{resume_final['total_ht_pa']:.2f} EUR"); st.metric(label="Total TTC Final (base P.A.)", value=f"{resume_final['total_ttc_pa']:.2f} EUR")
            with col2:
                st.metric(label="Total HT Final (base D.P.A.)", value=f"{resume_final['total_ht_dpa']:.2f} EUR"); st.metric(label="Total TTC Final (base D.P.A.)", value=f"{resume_final['total_ttc_dpa']:.2f} EUR")
            st.subheader("Exporter les résultats")
            excel_buffer = exporter_vers_excel_buffer_std(st.session_state.std_produits_finaux, st.session_state.std_resume_final, st.session_state.std_entreprise, st.session_state.std_marge, st.session_state.std_arrondi)
            date_str = datetime.now().strftime("%Y-%m-%d")
            nom_saisi = st.session_state.get("std_nom_calcul", "").strip(); nom_base_fichier = nom_saisi if nom_saisi else "Proposition_Tarifaire"
            nom_base_fichier_safe = re.sub(r'[\\/*?:"<>|]', "", nom_base_fichier); nom_fichier = f"{nom_base_fichier_safe}_{date_str}.xlsx"
            st.download_button(label="Telecharger le rapport standard", data=excel_buffer, file_name=nom_fichier, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_std")

with tab2:
    if st.session_state.df_base_clean is None: st.info("Veuillez charger un fichier Excel pour commencer.")
    else:
        df_base_clean = st.session_state.df_base_clean
        with st.expander("⚙️ Étape A: Configuration de la Promotion", expanded=True):
            st.text_input("Nom de la promotion (pour le nom du fichier)", key="promo_nom", placeholder="Ex: Promo Printemps 2024")
            col1, col2 = st.columns(2)
            with col1:
                if 'nom_fournisseur' in df_base_clean.columns:
                    fournisseurs_promo = ["Tous"] + sorted(df_base_clean['nom_fournisseur'].dropna().unique().tolist())
                    sel_fournisseur_promo = st.selectbox("Filtrer par Nom Fournisseur", options=fournisseurs_promo, key="promo_fournisseur")
                else: st.session_state.promo_fournisseur = "Tous"
            with col2:
                if 'code_famille' in df_base_clean.columns:
                    familles_promo = ["Tous"] + sorted(df_base_clean['code_famille'].dropna().unique().tolist())
                    sel_famille_promo = st.selectbox("Filtrer par Code Famille", options=familles_promo, key="promo_famille")
                else: st.session_state.promo_famille = "Tous"
            all_references_promo = sorted(pd.concat([df_base_clean['reference_article'].dropna().astype(str), df_base_clean['af_reffourniss'].dropna().astype(str), df_base_clean['reference_mbt'].dropna().astype(str)]).unique())
            promo_references = st.multiselect("OU sélectionner par références spécifiques", options=all_references_promo, key="promo_selection")
            if st.button("Préparer la promotion", key="promo_btn_preparer"):
                df_filtered = df_base_clean.copy()
                has_filters = False
                if 'nom_fournisseur' in df_base_clean.columns and st.session_state.promo_fournisseur != "Tous":
                    df_filtered = df_filtered[df_filtered['nom_fournisseur'] == st.session_state.promo_fournisseur]
                    has_filters = True
                if 'code_famille' in df_base_clean.columns and st.session_state.promo_famille != "Tous":
                    df_filtered = df_filtered[df_filtered['code_famille'] == st.session_state.promo_famille]
                    has_filters = True
                if not has_filters:
                    df_filtered = pd.DataFrame() 
                df_refs = pd.DataFrame()
                if st.session_state.promo_selection:
                    df_refs = rechercher_produits(df_base_clean, st.session_state.promo_selection)
                produits_promo = pd.concat([df_filtered, df_refs]).drop_duplicates(subset=['reference_article']).copy()
                if produits_promo.empty: 
                    st.warning("Aucun produit ne correspond à votre sélection.")
                else:
                    produits_promo['quantite'] = 1; produits_promo['prix_achat_promo'] = pd.NA
                    st.session_state.promo_df_a_modifier = produits_promo
                    if 'promo_resultats' in st.session_state: del st.session_state.promo_resultats
        if 'promo_df_a_modifier' in st.session_state:
            st.header("Étape B: Saisir les coûts et quantités")
            cols_a_editer_promo = ['reference_article', 'designation_article', 'prix_d_achat', 'prix_achat_promo', 'quantite']
            st.session_state.promo_df_modifie = st.data_editor(st.session_state.promo_df_a_modifier[cols_a_editer_promo], disabled=['reference_article', 'designation_article', 'prix_d_achat'], num_rows="dynamic", key="promo_editor")
            st.header("Étape C: Choisir la méthode de calcul")
            st.radio("Méthode de calcul du prix de vente promotionnel:", ("Appliquer une marge", "Définir un prix de vente cible"), key="promo_methode")
            if st.session_state.promo_methode == "Appliquer une marge":
                col1, col2 = st.columns(2)
                with col1: st.number_input("Marge promotionnelle (%)", min_value=0.0, max_value=99.9, value=15.0, step=0.5, format="%.2f", key="promo_marge")
                with col2: st.checkbox("Appliquer l'arrondi (,90€)", key="promo_arrondi")
            else:
                st.info("Veuillez saisir les prix de vente cibles dans le tableau ci-dessous.")
                df_pour_cible = st.session_state.promo_df_modifie.copy()
                if 'pv_promo_ht' not in df_pour_cible.columns: df_pour_cible['pv_promo_ht'] = pd.NA
                st.session_state.promo_df_cible = st.data_editor(df_pour_cible, disabled=['reference_article', 'designation_article', 'prix_d_achat', 'prix_achat_promo', 'quantite'], num_rows="dynamic", key="promo_editor_cible")
            if st.button("🚀 Calculer et Analyser la Promotion", key="promo_btn_calculer"):
                df_a_calculer = st.session_state.promo_df_cible if st.session_state.promo_methode == "Définir un prix de vente cible" else st.session_state.promo_df_modifie
                df_complet = pd.merge(df_a_calculer, st.session_state.df_base_clean, on='reference_article', how='left', suffixes=('', '_original'))
                st.session_state.promo_resultats = calculer_et_analyser_promo(df_complet.copy())
        if 'promo_resultats' in st.session_state:
            st.header("Étape D: Résultats de la Promotion")
            df_res = st.session_state.promo_resultats
            st.dataframe(df_res[['reference_article', 'designation_article', 'quantite', 'prix_achat_promo', 'pv_promo_ht', 'marge_degagee_%', 'remise_valeur_ht_€', 'remise_%']])
            st.subheader("Exporter le rapport de promotion")
            excel_buffer_promo = exporter_promo_vers_excel(df_res, st.session_state.promo_nom, st.session_state.promo_methode)
            date_str = datetime.now().strftime("%Y-%m-%d")
            nom_saisi = st.session_state.get("promo_nom", "").strip(); nom_base_fichier = nom_saisi if nom_saisi else "Rapport_Promotion"
            nom_base_fichier_safe = re.sub(r'[\\/*?:"<>|]', "", nom_base_fichier); nom_fichier = f"{nom_base_fichier_safe}_{date_str}.xlsx"
            st.download_button(label="Telecharger le rapport de promotion", data=excel_buffer_promo, file_name=nom_fichier, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_promo")

with tab3:
    if st.session_state.df_base_clean is None: st.info("Veuillez charger un fichier Excel pour commencer.")
    else:
        df_base_clean = st.session_state.df_base_clean
        with st.expander("⚙️ Configuration du Comparateur", expanded=True):
            analyze_all = st.checkbox("Analyser tous les produits de la base (selon filtres ci-dessous)")
            st.write("--- Filtres ---")
            col1, col2 = st.columns(2)
            with col1:
                if 'nom_fournisseur' in df_base_clean.columns:
                    fournisseurs = ["Tous"] + sorted(df_base_clean['nom_fournisseur'].dropna().unique().tolist())
                    sel_fournisseur = st.selectbox("Filtrer par Nom Fournisseur", options=fournisseurs, key="comp_nom_fournisseur")
                else: st.session_state.comp_nom_fournisseur = "Tous"
            with col2:
                if 'code_famille' in df_base_clean.columns:
                    familles = ["Tous"] + sorted(df_base_clean['code_famille'].dropna().unique().tolist())
                    sel_famille = st.selectbox("Filtrer par Code Famille", options=familles, key="comp_code_famille")
                else: st.session_state.comp_code_famille = "Tous"
            
            all_references_comp = sorted(pd.concat([df_base_clean['reference_article'].dropna().astype(str), df_base_clean['af_reffourniss'].dropna().astype(str), df_base_clean['reference_mbt'].dropna().astype(str)]).unique())
            sel_references = st.multiselect("OU Filtrer par Références spécifiques", options=all_references_comp, key="comp_selection")
            
            if st.button("🚀 Lancer la Comparaison", key="comp_btn_lancer"):
                filters = {'references': st.session_state.get('comp_selection', []), 'nom_fournisseur': st.session_state.get('comp_nom_fournisseur', 'Tous'), 'code_famille': st.session_state.get('comp_code_famille', 'Tous'),}
                
                # Exécution des calculs
                moins_cher, plus_cher = run_comparison(df_base_clean, filters, analyze_all)
                full_list = run_full_list(df_base_clean, filters, analyze_all)
                
                # Stockage en session
                st.session_state.comp_moins_cher = moins_cher
                st.session_state.comp_plus_cher = plus_cher
                st.session_state.comp_full_list = full_list
                st.session_state.comp_filters = filters
                st.session_state.comp_analyze_all = analyze_all

        if 'comp_full_list' in st.session_state:
            st.header("Résultats de la Comparaison")
            
            # 1. TABLEAU RÉCAPITULATIF COMPLET (DEMANDE UTILISATEUR)
            st.subheader("📋 Récapitulatif Complet des Tarifs")
            df_recap = st.session_state.comp_full_list.copy()
            # Renommage des colonnes pour l'affichage
            df_recap_display = df_recap.rename(columns={
                'reference_article': 'Référence TTL',
                'reference_mbt': 'Référence MBT',
                'designation_article': 'Désignation',
                'prix_de_vente': 'Prix de vente TTL',
                'prix_de_vente_mbt': 'Prix de vente MBT'
            })
            st.dataframe(df_recap_display, use_container_width=True)
            
            st.divider()
            
            # 2. TABLEAUX D'ÉCARTS (MOINS CHER / PLUS CHER)
            cols_to_show = ['reference_article', 'designation_article', 'prix_de_vente', 'prix_de_vente_mbt', 'ecart_€', 'ecart_%', 'dernier_prix_d_achat', 'marge_ttl_%', 'marge_mbt_%', 'alertes']
            if not st.session_state.comp_moins_cher.empty:
                st.subheader("✅ MBT est Moins Cher que TTL")
                st.dataframe(st.session_state.comp_moins_cher[cols_to_show])
            
            if not st.session_state.comp_plus_cher.empty:
                st.subheader("❌ MBT est Plus Cher que TTL")
                st.dataframe(st.session_state.comp_plus_cher[cols_to_show])
            
            if st.session_state.comp_moins_cher.empty and st.session_state.comp_plus_cher.empty:
                st.info("Aucun écart de prix trouvé pour la sélection actuelle.")
            
            # EXPORT EXCEL
            st.subheader("Exporter le rapport")
            excel_buffer_comp = exporter_comparaison_vers_excel(
                st.session_state.comp_moins_cher, 
                st.session_state.comp_plus_cher, 
                st.session_state.comp_full_list,
                st.session_state.comp_filters, 
                st.session_state.comp_analyze_all
            )
            date_str = datetime.now().strftime("%Y-%m-%d"); nom_fichier = f"Comparaison_TTL_MBT_{date_str}.xlsx"
            st.download_button(label="Telecharger le rapport Excel (3 onglets)", data=excel_buffer_comp, file_name=nom_fichier, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_comp")

with tab4:
    if st.session_state.df_base_clean is None: st.info("Veuillez charger un fichier Excel pour commencer.")
    else:
        df_base_clean = st.session_state.df_base_clean
        with st.expander("⚙️ Configuration du Vérificateur de Marges", expanded=True):
            seuil_marge = st.number_input("Seuil de marge minimum (%)", min_value=0.1, max_value=100.0, value=20.0, step=1.0, key="margin_threshold")
            analyze_all_margins = st.checkbox("Analyser tous les produits")
            st.write("--- OU ---")
            col1, col2 = st.columns(2)
            with col1:
                if 'code_famille' in df_base_clean.columns:
                    familles = ["Tous"] + sorted(df_base_clean['code_famille'].dropna().unique().tolist())
                    sel_famille_margin = st.selectbox("Filtrer par Code Famille", options=familles, key="margin_famille")
                else: st.session_state.margin_famille = "Tous"
            with col2:
                if 'nom_fournisseur' in df_base_clean.columns:
                    fournisseurs = ["Tous"] + sorted(df_base_clean['nom_fournisseur'].dropna().unique().tolist())
                    sel_fournisseur_margin = st.selectbox("Filtrer par Nom Fournisseur", options=fournisseurs, key="margin_fournisseur")
                else: st.session_state.margin_fournisseur = "Tous"
            all_references_margin = sorted(pd.concat([df_base_clean['reference_article'].dropna().astype(str), df_base_clean['af_reffourniss'].dropna().astype(str), df_base_clean['reference_mbt'].dropna().astype(str)]).unique())
            sel_references_margin = st.multiselect("Filtrer par Références", options=all_references_margin, key="margin_selection")
            if st.button("🚀 Lancer la Vérification des Marges", key="margin_btn_lancer"):
                filters = {'references': st.session_state.get('margin_selection', []), 'nom_fournisseur': st.session_state.get('margin_fournisseur', 'Tous'), 'code_famille': st.session_state.get('margin_famille', 'Tous'),}
                alerts_ttl, alerts_mbt = run_margin_check(df_base_clean, st.session_state.margin_threshold, filters, analyze_all_margins)
                st.session_state.margin_alerts_ttl = alerts_ttl; st.session_state.margin_alerts_mbt = alerts_mbt
        if 'margin_alerts_ttl' in st.session_state:
            st.header("Résultats de la Vérification")
            if not st.session_state.margin_alerts_ttl.empty:
                st.subheader("📉 Alertes de Marge Insuffisante - TTL")
                st.dataframe(st.session_state.margin_alerts_ttl[['reference_article', 'designation_article', 'dernier_prix_d_achat', 'prix_de_vente', 'marge_ttl_%', 'alerte_remise']])
            else:
                st.info("Aucune alerte de marge trouvée pour TTL.")
            if not st.session_state.margin_alerts_mbt.empty:
                st.subheader("📉 Alertes de Marge Insuffisante - MBT")
                st.dataframe(st.session_state.margin_alerts_mbt[['reference_article', 'designation_article', 'dernier_prix_d_achat', 'prix_de_vente_mbt', 'marge_mbt_%', 'alerte_remise']])
            else:
                st.info("Aucune alerte de marge trouvée pour MBT.")
            st.subheader("Exporter le rapport des marges")
            excel_buffer_margin = exporter_marges_vers_excel(st.session_state.margin_alerts_ttl, st.session_state.margin_alerts_mbt, st.session_state.margin_threshold)
            date_str = datetime.now().strftime("%Y-%m-%d"); nom_fichier = f"Rapport_Marges_{date_str}.xlsx"
            st.download_button(label="Telecharger le rapport des marges", data=excel_buffer_margin, file_name=nom_fichier, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_margin")

with tab5:
    if st.session_state.df_base_clean is None: st.info("Veuillez charger un fichier Excel pour commencer.")
    else:
        df_base_clean = st.session_state.df_base_clean
        with st.expander("📦 Étape 1 : Configuration du Bundle", expanded=True):
            st.text_input("Nom de la Nomenclature / Bundle", key="nom_bundle", placeholder="Ex: Pack Studio Débutant")
            all_references_nom = sorted(pd.concat([df_base_clean['reference_article'].dropna().astype(str), df_base_clean['af_reffourniss'].dropna().astype(str), df_base_clean['reference_mbt'].dropna().astype(str)]).unique())
            references_bundle = st.multiselect("Sélectionnez les produits composants", options=all_references_nom, key="bundle_selection")
            
            if st.button("Préparer la liste des composants", key="btn_prep_bundle"):
                if not references_bundle: st.warning("Veuillez sélectionner au moins un produit.")
                else:
                    composants = rechercher_produits(df_base_clean, references_bundle).copy()
                    composants['quantite'] = 1
                    st.session_state.bundle_df_a_modifier = composants
                    if 'bundle_res_final' in st.session_state: del st.session_state.bundle_res_final

        if 'bundle_df_a_modifier' in st.session_state:
            st.header("Étape 2 : Quantités et Prix d'achat")
            cols_edit_bundle = ['reference_article', 'designation_article', 'prix_d_achat', 'prix_de_vente', 'prix_de_vente_mbt', 'quantite']
            st.session_state.bundle_df_modifie = st.data_editor(st.session_state.bundle_df_a_modifier[cols_edit_bundle], 
                                                                disabled=['reference_article', 'designation_article', 'prix_de_vente', 'prix_de_vente_mbt'], 
                                                                num_rows="dynamic", key="bundle_editor")
            
            st.header("Étape 3 : Saisie des paramètres de simulation")
            col_b1, col_b2 = st.columns(2)
            with col_b1:
                v_pv_cible = st.number_input("Simulation 1 : Prix de vente Bundle HT cible (€)", min_value=0.0, value=100.0, step=10.0, key="inp_pv_cible")
            with col_b2:
                v_marge_cible = st.number_input("Simulation 2 : Marge à appliquer sur le total (%)", min_value=0.0, max_value=99.0, value=30.0, step=1.0, key="inp_marge_cible")
                arrondi_bundle = st.checkbox("Appliquer l'arrondi à X,90€ (uniquement pour Sim 2)", value=True, key="inp_arrondi_bundle")
            
            if st.button("🚀 Calculer les Simulations", key="btn_calcul_bundle"):
                st.session_state.bundle_res_final = calculer_simulation_nomenclature(st.session_state.bundle_df_modifie, v_pv_cible, v_marge_cible, arrondi_bundle)

        if 'bundle_res_final' in st.session_state:
            st.header("Étape 4 : Résultats des Simulations")
            res = st.session_state.bundle_res_final
            st.subheader("Détail des composants")
            st.dataframe(res['df'][['reference_article', 'designation_article', 'quantite', 'prix_d_achat', 'prix_de_vente', 'prix_de_vente_mbt', 'total_pa_ligne']])
            
            col_res1, col_res2 = st.columns(2)
            with col_res1:
                st.markdown("### 🎯 Simulation 1 : Prix Cible")
                st.metric("Prix de Vente (HT)", f"{res['sim_prix']['pv_ht']:.2f} €")
                st.metric("Prix de Vente (TTC)", f"{res['sim_prix']['pv_ttc']:.2f} €")
                st.metric("Marge dégagée", f"{res['sim_prix']['marge']*100:.2f} %")
            with col_res2:
                st.markdown("### 📈 Simulation 2 : Marge Cible")
                st.metric("Prix de Vente (HT)", f"{res['sim_marge']['pv_ht']:.2f} €")
                st.metric("Prix de Vente (TTC)", f"{res['sim_marge']['pv_ttc']:.2f} €")
                st.metric("Marge (Appliquée)", f"{res['sim_marge']['marge']*100:.2f} %")
            
            st.divider()
            c_norm1, c_norm2, c_norm3 = st.columns(3)
            c_norm1.metric("Total Coût Achat Bundle (HT)", f"{res['cout_total_pa']:.2f} €")
            c_norm2.metric("Total Normal Tetenal (HT)", f"{res['total_normal_ttl']:.2f} €")
            c_norm3.metric("Total Normal MB TECH (HT)", f"{res['total_normal_mbt']:.2f} €")
            
            st.subheader("Exporter la Nomenclature Comparative")
            buf_bundle = exporter_nomenclature_comparative_excel(res, st.session_state.nom_bundle)
            nom_f = f"Nomenclature_{re.sub(r'[\\/*?::<>|]', '', st.session_state.nom_bundle)}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            st.download_button("Télécharger le rapport Excel complet", data=buf_bundle, file_name=nom_f, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
